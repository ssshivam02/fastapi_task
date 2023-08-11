from fastapi import APIRouter, Request
import pandas as pd
import json
import openpyxl
# import pytest
from excel import create_excelsheet
router = APIRouter()

@router.get("/get_list_employee_details")
async def get_list_employees():
    # df = pd.read_excel("employee_attendace.xlsx")
    # print("\n")
    # print(df)
    # print(df[df["E.Name"]== "Shivam"])
    # # return df
    work_book = openpyxl.load_workbook("employee_attendace.xlsx")
    sheet_obj = work_book["employee_data"]
    res = {}
    for row in sheet_obj.iter_rows(min_row=2,values_only= True):
        S_No = row[0]
        employee_detail= {
            "E.Code": row[1], 
            "E.Name" : row[2], 
            "Date": row[3],
            "Attendance":  row[4]
        }
        res[S_No] = employee_detail
    
    return json.dumps(res, indent=4)

@router.post("/add_employee")
async def add_employee_detail(request: Request):
    req_data = await request.json()
    create_excelsheet([req_data])
    work_book = openpyxl.load_workbook("employee_attendace.xlsx")
    sheet_obj = work_book["employee_data"]
    res = {}
    max_row = sheet_obj.max_row
    for row in sheet_obj.iter_rows(min_row=max_row,values_only= True):
        S_No = row[0]
        employee_detail= {
            "E.Code": row[1], 
            "E.Name" : row[2], 
            "Date": row[3],
            "Attendance":  row[4]
        }
        res[S_No] = employee_detail
    
    return json.dumps(res, indent=4)

@router.get("/employee_detail/{employee_code}")
async def get_employee_detail(employee_code):
    work_book = openpyxl.load_workbook("employee_attendace.xlsx")
    sheet_obj = work_book["employee_data"]
    res = {}
    for row in sheet_obj.iter_rows(min_row=2,values_only= True):
        if employee_code == row[1]:
            S_No = row[0]
            employee_detail= {
                "E.Code": row[1], 
                "E.Name" : row[2], 
                "Date": row[3],
                "Attendance":  row[4]
            }
            res[S_No] = employee_detail

    return json.dumps(res, indent=4)

@router.get("/employee_detail/absent/list")
async def get_employee_detail_absent():
    work_book = openpyxl.load_workbook("employee_attendace.xlsx")
    sheet_obj = work_book["employee_data"]
    res = {}
    for row in sheet_obj.iter_rows(min_row=2,values_only= True):
        if row[1] not in res.keys():
            if row[4]=="A":
                employee_detail= [{ 
                    "E.Name" : row[2], 
                    "Date": row[3],
                    "Attendance":  row[4]
                }]
                res[row[1]] = employee_detail
        else:
            if row[4]=="A":
                employee_detail= { 
                    "E.Name" : row[2], 
                    "Date": row[3],
                    "Attendance":  row[4]
                }
                res[row[1]].append(employee_detail)  
    # import pprint
    # pprint.pprint(res)    
    return json.dumps(res, indent=4)

@router.get("/send_email")
async def get_send_email():
    work_book = openpyxl.load_workbook("employee_attendace.xlsx")
    sheet_obj = work_book["employee_data"]
    res = {}
    for row in sheet_obj.iter_rows(min_row=2,values_only= True):
        if row[1] not in res.keys():
            if row[4]=="A":
                employee_detail= [{ 
                    "E.Name" : row[2], 
                    "Date": row[3],
                    "Attendance":  row[4]
                }]
                res[row[1]] = employee_detail
        else:
            if row[4]=="A":
                employee_detail= { 
                    "E.Name" : row[2], 
                    "Date": row[3],
                    "Attendance":  row[4]
                }
                res[row[1]].append(employee_detail)  
    for emp_code in res:
        if len(res[emp_code]) ==2:
            days = len(res[emp_code])
            # employee mail body
            name = res[emp_code][0]["E.Name"]
            message = f"""
                Hi {name},
                This email is to inform you that according to our attendance records, you are absent from your duties
                for {days} days. Please apply for leave
                Thanks
                HR Team
            """
            # import smtplib

            # # creates SMTP session
            # s = smtplib.SMTP('smtp.gmail.com', 587)

            # # start TLS for security
            # s.starttls()

            # # Authentication
            # s.login("sender_email_id", "sender_email_id_password")

            # # message to be sent

            # # sending the mail
            # s.sendmail("sender_email_id", "receiver_email_id", message)

            # # terminating the session
            # s.quit()
            print(message)

        elif len(res[emp_code])>=3:
            # Employer
            days = len(res[emp_code])
            name = res[emp_code][0]["E.Name"]
            message = f"""
                Hi,
                This email is to inform you that according to attendance records, {name} did not attend to the duties
                for {days} days.
                Thanks
                HR Team
                """
            # import smtplib

            # # creates SMTP session
            # s = smtplib.SMTP('smtp.gmail.com', 587)

            # # start TLS for security
            # s.starttls()

            # # Authentication
            # s.login("sender_email_id", "sender_email_id_password")

            # # message to be sent

            # # sending the mail
            # s.sendmail("sender_email_id", "receiver_email_id", message)

            # # terminating the session
            # s.quit()
            print(message)


