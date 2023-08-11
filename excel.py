import openpyxl
import os 
from openpyxl.styles import Font, Alignment
assessment_data = [{
    "E.Code": "100", "E.Name" : "John", "Date": "01-08-2023","Attendance": "P"},{
    "E.Code": "101", "E.Name" : "Chandu", "Date": "01-08-2023","Attendance": "A"},{
    "E.Code": "102", "E.Name": "Jessi", "Date": "01-08-2023","Attendance": "L"},{
    "E.Code": "103", "E.Name" : "Sunny", "Date": "01-08-2023","Attendance": "P"},{
    "E.Code": "104", "E.Name" : "Shivam", "Date": "01-08-2023","Attendance": "A"}
    ]
def create_excelsheet(assessment_data):
    try:
        # Create a new workbook using openpyxl
        path = os.getcwd() + "/employee_attendace.xlsx"

        if not os.path.exists(path):
            wb = openpyxl.Workbook()
            # Assign the active sheet to ws_employee
            ws_employee = wb.active
            # Set the title of the worksheet to "employee"
            ws_employee.title = "employee_data"

            # Create and append the summary data to the worksheet
            summary_data = ("S.No", "E.Code", "E.Name", "Date", "Attendance")
            ws_employee.append(summary_data)

            # Set the font style of the first row to bold
            ft = Font(bold=True)
            for row in ws_employee["A1:E1"]:
                for cell in row:
                    cell.font = ft
                    cell.alignment = Alignment(horizontal= "center", vertical= "center")
            file_path = os.getcwd() + '/employee_attendace.xlsx'
            wb.save(file_path)

        wb = openpyxl.load_workbook("employee_attendace.xlsx")
        ws_employee = wb.active
        print(assessment_data)
        # Append Employee assessment data under the respective columns
        for data in assessment_data:
            print(41, assessment_data)
            max_row = ws_employee.max_row
            summary_data = (max_row, data["E.Code"], data["E.Name"], data["Date"], data["Attendance"])
            ws_employee.append(summary_data)

        # Save the workbook as "employee_attendace.xlsx" in the current working directory
            file_path = os.getcwd() + '/employee_attendace.xlsx'
            wb.save(file_path)
    except Exception as err:
        # Raise a custom CommonError with error message generated from get_file_info
        print(err)

if __name__ == "__main__":
    print(create_excelsheet(assessment_data))

